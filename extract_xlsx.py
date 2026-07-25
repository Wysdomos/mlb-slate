"""
extract_xlsx.py -- Daily MLB Slate Excel -> JSON extractor
Usage:  python3 extract_xlsx.py <path_to_xlsx> [output_json]
"""

import sys
import json
import os
import glob
import re
import datetime as _dt
from openpyxl import load_workbook

PROJECTED_SENTINEL = "__PROJECTED_MODE__"

def _wb_date(path):
    """Parse the date embedded in a slate filename, e.g. 'MLB_Slate_6-16-26.xlsx' -> 2026-06-16."""
    m = re.search(r'(\d{1,2})[-_ ](\d{1,2})[-_ ](\d{2,4})', os.path.basename(path))
    if not m:
        return None
    mo, d, y = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    try:
        return _dt.date(y, mo, d)
    except ValueError:
        return None

def find_xlsx():
    # An explicit .xlsx path on the command line always wins.
    for a in sys.argv[1:]:
        if a.lower().endswith('.xlsx'):
            if not os.path.exists(a):
                print(f"ERROR: File not found: {a}", file=sys.stderr)
                sys.exit(1)
            return a
    matches = [f for f in (glob.glob("*.xlsx") + glob.glob("**/*.xlsx", recursive=False))
               if not os.path.basename(f).startswith('~$')]
    if not matches:
        if os.environ.get("ALLOW_PROJECTED_MODE") == "1":
            return None
        print("ERROR: No .xlsx file found.", file=sys.stderr)
        sys.exit(1)
    # Pick the workbook whose filename date is newest (today's upload), not the alphabetical first.
    chosen = max(matches, key=lambda f: (_wb_date(f) or _dt.date.min, os.path.getmtime(f)))
    if len(matches) > 1:
        print(f"Multiple xlsx found ({len(matches)}); using newest by date: {chosen}", file=sys.stderr)
    return chosen

def resolve_output():
    for a in sys.argv[1:]:
        if a.lower().endswith('.json'):
            return a
    return "day_data.json"

def projected_marker():
    today = _today_et()
    keys = [
        "HR_Leaderboard",
        "Hit_Probabilities",
        "Sweet_Spot_Analyzer",
        "Pitcher_Projections",
        "SP_Projections",
        "Park_Factors",
        "Sweet_Spot_Slate",
        "BP_Batters",
        "BP_Pitchers",
        "BP_Teams",
        "BP_Games",
        "Streaks",
        "Scout",
        "Best_Spots",
    ]
    data = {key: [] for key in keys}
    data["_mode"] = "projected"
    data["_slate_date"] = today.isoformat()
    data["INDEX"] = [{"Sheet": key, "Rows": 0, "Cols": 0} for key in keys]
    return data

def sheet_to_rows(ws):
    headers = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            if any(c is not None for c in row):
                headers = [str(c).strip() if c is not None else f"Col{i}"
                           for i, c in enumerate(row)]
            continue
        if not any(c is not None for c in row):
            continue
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                record[headers[i]] = val
        rows.append(record)
    return rows

def clean_val(v):
    if v is None:
        return None
    if isinstance(v, float):
        if v != v:
            return None
        if v == int(v):
            return int(v)
    return v

def clean_rows(rows):
    return [{k: clean_val(v) for k, v in r.items()} for r in rows]

def extract(xlsx_path):
    print(f"Reading: {xlsx_path}", file=sys.stderr)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}", file=sys.stderr)

    SHEET_MAP = {
        "HR_Leaderboard":      "HR_Leaderboard",
        "Hit_Probabilities":   "Hit_Probabilities",
        "Sweet_Spot_Analyzer": "Sweet_Spot_Analyzer",
        "Pitcher_Projections": "Pitcher_Projections",
        "SP_Projections":      "SP_Projections",
        "Park_Factors":        "Park_Factors",
        "Sweet_Spot_Slate":    "Sweet_Spot_Slate",
        "BP_Batters":          "BP_Batters",
        "BP_Pitchers":         "BP_Pitchers",
        "BP_Teams":            "BP_Teams",
        "BP_Games":            "BP_Games",
        "Streaks":             "Streaks",
        "Scout":              "Scout",
        "Best_Spots":         "Best_Spots",
    }

    data = {}
    index_rows = []

    for sheet_name, json_key in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found -- skipping", file=sys.stderr)
            data[json_key] = []
            continue
        ws = wb[sheet_name]
        rows = clean_rows(sheet_to_rows(ws))
        data[json_key] = rows
        cols = len(rows[0]) if rows else 0
        index_rows.append({"Sheet": sheet_name, "Rows": len(rows), "Cols": cols})
        print(f"  {sheet_name}: {len(rows)} rows", file=sys.stderr)

    data["INDEX"] = index_rows
    wb.close()
    return data

def _today_et():
    """Real calendar date in US/Eastern -- the timezone MLB slates run on."""
    from zoneinfo import ZoneInfo
    return _dt.datetime.now(ZoneInfo("America/New_York")).date()

if __name__ == "__main__":
    if '--which' in sys.argv:
        chosen = find_xlsx()
        if chosen is None:
            print(PROJECTED_SENTINEL)
            sys.exit(0)
        wb_date, today = _wb_date(chosen), _today_et()
        if wb_date != today:
            print(f"ERROR: newest available slate file is dated {wb_date} "
                  f"but today (ET) is {today} -- no fresh upload found. "
                  f"Refusing to build a stale slate.", file=sys.stderr)
            sys.exit(1)
        print(chosen)
        sys.exit(0)
    xlsx_path = find_xlsx()
    output_path = resolve_output()
    if xlsx_path is None:
        data = projected_marker()
        print(
            f"Projected Mode marker written for {data['_slate_date']} "
            f"because no workbook was uploaded and ALLOW_PROJECTED_MODE=1.",
            file=sys.stderr,
        )
    else:
        wb_date, today = _wb_date(xlsx_path), _today_et()
        if wb_date != today:
            print(f"ERROR: slate file is dated {wb_date} "
                  f"but today (ET) is {today} -- refusing stale workbook.",
                  file=sys.stderr)
            sys.exit(1)
        data = extract(xlsx_path)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    total_rows = sum(len(v) for v in data.values() if isinstance(v, list))
    print(f"Done. {total_rows} total rows -> {output_path}", file=sys.stderr)
