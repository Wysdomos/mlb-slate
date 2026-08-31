"""
grade_results.py — For The Record grader.

Reads yesterday's committed slate_picks.json and grades each call:
  * HR  — primary source is the workbook's HR_Results_<date> tab (no API needed);
          win if the batter homered.
  * K / Hits / HRR / SB / 2B / Totals / NRFI — graded from MLB Stats API box scores
          when reachable (GitHub Actions). When the API is unavailable (e.g. sandbox),
          those markets are left 'pending' rather than guessed.

Accumulates a per-day history in results.json so the season record, last-7 and trend grow
each morning. Always exits 0 so a grading hiccup never blocks the slate build.

Env:
  PICKS_FILE     slate_picks.json (default)
  RESULTS_XLSX   workbook containing the HR_Results_<date> tab (optional)
  RESULTS_FILE   results.json output (default)
  GRADE_DATE     label for the graded day (optional)
"""
import json, os, re, sys, unicodedata

PICKS_FILE   = os.environ.get('PICKS_FILE', 'slate_picks.json')
RESULTS_XLSX = os.environ.get('RESULTS_XLSX', '')
RESULTS_FILE = os.environ.get('RESULTS_FILE', 'results.json')
GRADE_DATE   = os.environ.get('GRADE_DATE', '')

def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z ]', '', s.lower()).strip()

def load_json(path, default):
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return default

def read_hr_results(xlsx):
    """Return ({homering batters}, date_label) from the workbook's HR_Results_<date> tab."""
    if not xlsx or not os.path.exists(xlsx):
        return set(), ''
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        tab = next((s for s in wb.sheetnames if s.lower().startswith('hr_results')), None)
        if not tab:
            return set(), ''
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c else '' for c in rows[0]]
        bi = hdr.index('Batter') if 'Batter' in hdr else 0
        homers = {norm(r[bi]) for r in rows[1:] if r and r[bi]}
        date = tab.replace('HR_Results_', '').replace('HR_Results', '')
        return homers, date
    except Exception as e:
        print('HR results read failed:', e)
        return set(), ''

BDL_KEY = (os.environ.get('BDL_KEY') or os.environ.get('BALLDONTLIE_API_KEY') or '').strip()
_DBG = {'bdl': False}

def _g(d, *keys, default=0):
    """First present key — handles MLB field-name variants across providers."""
    for k in keys:
        if isinstance(d, dict) and d.get(k) is not None:
            return d[k]
    return default

def fetch_bdl(date, key):
    """balldontlie MLB box scores (your GOAT key) -> per-player actuals.
    GET /mlb/v1/box_scores?date=YYYY-MM-DD  with header  Authorization: <key>."""
    try:
        import urllib.request
        req = urllib.request.Request(
            f'https://api.balldontlie.io/mlb/v1/box_scores?date={date}',
            headers={'Authorization': key})
        payload = json.load(urllib.request.urlopen(req, timeout=12))
        out = {'batters': {}, 'pitchers': {}, 'first_inning': {}, 'totals': {}}
        games = payload.get('data', [])
        for g in games:
            for side in ('home_team', 'visitor_team'):
                for e in (g.get(side, {}) or {}).get('players', []):
                    pl = e.get('player', {}) or {}
                    nm = norm(f"{pl.get('first_name','')} {pl.get('last_name','')}")
                    if not nm:
                        continue
                    if not _DBG['bdl']:   # one-time schema dump -> reveals real field names in the Actions log
                        print('BDL sample player keys:', sorted(e.keys())); _DBG['bdl'] = True
                    # batting (defensive field names)
                    if _g(e, 'at_bats', 'ab', 'plate_appearances', 'hits', 'h', default=None) is not None:
                        out['batters'][nm] = {
                            'h': _g(e, 'hits', 'h'), 'hr': _g(e, 'home_runs', 'homeruns', 'hr'),
                            'r': _g(e, 'runs', 'runs_scored', 'r'), 'rbi': _g(e, 'rbi', 'runs_batted_in', 'rbis'),
                            'sb': _g(e, 'stolen_bases', 'sb'), 'd': _g(e, 'doubles', 'double', '2b'),
                            't': _g(e, 'triples', 'triple', '3b')}
                    # pitching — strikeouts decide the K market; rest is context
                    k_ = _g(e, 'strikeouts_pitching', 'pitching_strikeouts', 'strike_outs', 'strikeouts', 'so', default=None)
                    if k_ is not None and (e.get('innings_pitched') or e.get('ip') or e.get('batters_faced')):
                        out['pitchers'][nm] = {
                            'k': k_ or 0, 'h': _g(e, 'hits_allowed', 'pitching_hits', 'hits_against'),
                            'er': _g(e, 'earned_runs', 'er'), 'outs': _g(e, 'outs', 'outs_pitched')}
        print(f'balldontlie: {len(games)} games -> {len(out["batters"])} batters, {len(out["pitchers"])} pitchers')
        return out
    except Exception as e:
        print('balldontlie unavailable:', e)
        return {}

def fetch_box_results(date):
    """MLB Stats API box scores (free, no key) -> per-player actuals. Fallback source."""
    try:
        import urllib.request
        base = 'https://statsapi.mlb.com/api/v1'
        sched = json.load(urllib.request.urlopen(
            f'{base}/schedule?sportId=1&date={date}&hydrate=linescore', timeout=8))
        out = {'batters': {}, 'pitchers': {}, 'first_inning': {}, 'totals': {}}
        for d in sched.get('dates', []):
            for g in d.get('games', []):
                if g.get('status', {}).get('abstractGameState') != 'Final':
                    continue
                pk = g['gamePk']
                box = json.load(urllib.request.urlopen(f'{base}/game/{pk}/boxscore', timeout=8))
                for side in ('home', 'away'):
                    team = box['teams'][side]
                    for pid, p in team.get('players', {}).items():
                        nm = norm(p.get('person', {}).get('fullName', ''))
                        bs = p.get('stats', {}).get('batting', {})
                        ps = p.get('stats', {}).get('pitching', {})
                        if bs:
                            out['batters'][nm] = {
                                'h': bs.get('hits', 0), 'hr': bs.get('homeRuns', 0),
                                'r': bs.get('runs', 0), 'rbi': bs.get('rbi', 0),
                                'sb': bs.get('stolenBases', 0), 'd': bs.get('doubles', 0),
                                't': bs.get('triples', 0)}
                        if ps:
                            out['pitchers'][nm] = {
                                'k': ps.get('strikeOuts', 0), 'h': ps.get('hits', 0),
                                'er': ps.get('earnedRuns', 0), 'outs': ps.get('outs', 0)}
        return out
    except Exception as e:
        print('Box-score API unavailable (markets stay pending):', e)
        return {}

def _wb_date(path):
    import re, datetime as _dt
    m = re.search(r'(\d{1,2})[-_ ](\d{1,2})[-_ ](\d{2,4})', os.path.basename(path))
    if not m:
        return None
    mo, d, y = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    try:
        return _dt.date(y, mo, d)
    except ValueError:
        return None

def find_workbook():
    if RESULTS_XLSX and os.path.exists(RESULTS_XLSX):
        return RESULTS_XLSX
    import glob, datetime as _dt
    cands = [f for f in glob.glob('*.xlsx')
             if 'slate' in os.path.basename(f).lower() and not os.path.basename(f).startswith('~$')]
    if not cands:
        return ''
    # newest workbook (today's upload holds yesterday's HR_Results tab)
    return max(cands, key=lambda f: (_wb_date(f) or _dt.date.min, os.path.getmtime(f)))

_TEAM = {'AZ':'ARI','CHW':'CWS','KCR':'KC','SDP':'SD','SFG':'SF','TBR':'TB','ATH':'OAK',
         'WAS':'WSH','WSN':'WSH','LV':'OAK'}
def canon(ab):
    ab = (ab or '').strip().upper()
    return _TEAM.get(ab, ab)
def gamekey(game):
    """'KC@WAS' -> 'KC@WSH' (canonical away@home)."""
    if '@' in (game or ''):
        a, h = game.split('@', 1)
        return f'{canon(a)}@{canon(h)}'
    return (game or '').upper()

def fetch_games(date):
    """Game-level results (final total runs + first-inning runs) from the free MLB Stats API
    linescore. Drives Totals and NRFI regardless of the player-prop source. Keyed away@home."""
    out = {'totals': {}, 'first_inning': {}}
    try:
        import urllib.request
        base = 'https://statsapi.mlb.com/api/v1'
        sched = json.load(urllib.request.urlopen(
            f'{base}/schedule?sportId=1&date={date}&hydrate=linescore,team', timeout=10))
        for d in sched.get('dates', []):
            for g in d.get('games', []):
                if g.get('status', {}).get('abstractGameState') != 'Final':
                    continue
                t = g.get('teams', {})
                ah = (t.get('away', {}).get('team', {}) or {}).get('abbreviation')
                hh = (t.get('home', {}).get('team', {}) or {}).get('abbreviation')
                if not ah or not hh:
                    continue
                key = f'{canon(ah)}@{canon(hh)}'
                ls = g.get('linescore', {}) or {}
                ar = t.get('away', {}).get('score'); hrn = t.get('home', {}).get('score')
                if ar is None or hrn is None:
                    lt = ls.get('teams', {}) or {}
                    ar = (lt.get('away', {}) or {}).get('runs'); hrn = (lt.get('home', {}) or {}).get('runs')
                if ar is not None and hrn is not None:
                    out['totals'][key] = ar + hrn
                innings = ls.get('innings', []) or []
                if innings:
                    fi = innings[0]
                    out['first_inning'][key] = ((fi.get('away', {}) or {}).get('runs', 0) or 0) \
                                             + ((fi.get('home', {}) or {}).get('runs', 0) or 0)
        print(f'games: {len(out["totals"])} totals, {len(out["first_inning"])} first-inning; '
              f'keys={sorted(out["totals"].keys())}')
    except Exception as e:
        print('game-level API unavailable (Totals/NRFI pending):', e)
    return out

def to_iso(hr_date, iso_date):
    """Box-score API needs YYYY-MM-DD. Prefer the archived slate_date; else build from the tab suffix."""
    if iso_date:
        return iso_date
    if hr_date and '-' in hr_date:
        import datetime as _dt
        try:
            mm, dd = hr_date.split('-')[:2]
            return f'{_dt.date.today().year}-{int(mm):02d}-{int(dd):02d}'
        except Exception:
            pass
    return None

def choose_slate(hr_date):
    """Which day's picks to grade. Priority: explicit GRADE_DATE -> workbook results tab ->
    most recent COMMITTED archive that finished before today (ET) -> yesterday fallback.
    Returns (picks_path, md_label, iso_date)."""
    import glob, datetime as _dt
    if GRADE_DATE and os.path.exists(f'slate_picks_{GRADE_DATE}.json'):
        f = f'slate_picks_{GRADE_DATE}.json'
        return f, GRADE_DATE, load_json(f, {}).get('slate_date')
    if hr_date and os.path.exists(f'slate_picks_{hr_date}.json'):
        f = f'slate_picks_{hr_date}.json'
        return f, hr_date, load_json(f, {}).get('slate_date')
    today_et = (_dt.datetime.now(_dt.timezone.utc) - _dt.timedelta(hours=4)).date().isoformat()
    best = best_iso = None
    for f in glob.glob('slate_picks_*.json'):
        sd = load_json(f, {}).get('slate_date')
        if sd and sd < today_et and (best_iso is None or sd > best_iso):
            best_iso, best = sd, f
    if best:
        try:
            _, m, d = best_iso.split('-'); md = f'{int(m)}-{int(d)}'
        except Exception:
            md = best_iso
        return best, md, best_iso
    yest = (_dt.datetime.now(_dt.timezone.utc) - _dt.timedelta(hours=4) - _dt.timedelta(days=1)).date()
    return PICKS_FILE, f'{yest.month}-{yest.day}', yest.isoformat()

def grade():
    xlsx = find_workbook()
    homers, hr_date = read_hr_results(xlsx)
    picks_path, date, iso_date = choose_slate(hr_date)
    payload = load_json(picks_path, {})
    picks = payload.get('picks', [])
    if not iso_date:
        iso_date = payload.get('slate_date')
    api_date = to_iso(date, iso_date)            # full date for the API
    src = 'balldontlie' if BDL_KEY else 'MLB Stats API'
    print(f'Grading {date} (API date {api_date}, source {src}): {len(picks)} picks from {picks_path}, '
          f'{len(homers)} homers from {xlsx or "no workbook"}')
    box = {}
    if api_date:
        if BDL_KEY:
            box = fetch_bdl(api_date, BDL_KEY)
            if not box.get('batters') and not box.get('pitchers'):
                print('balldontlie returned no players — falling back to MLB Stats API')
                box = fetch_box_results(api_date)
        else:
            box = fetch_box_results(api_date)
    if not isinstance(box, dict):
        box = {}
    for k in ('batters', 'pitchers', 'totals', 'first_inning'):
        box.setdefault(k, {})
    if api_date:
        games = fetch_games(api_date)
        box['totals'].update(games.get('totals', {}))
        box['first_inning'].update(games.get('first_inning', {}))

    graded = []   # {market, name, line, consensus, win(bool/None), got, detail}
    for p in picks:
        mkt = p['market']; nm = norm(p['name']) if p.get('name') else ''
        win, got = None, '—'
        detail = {}
        def directional(actual):
            direction = (p.get('direction') or 'Over').lower()
            target = p.get('win_at')
            if target is None:
                return None
            return actual <= target if direction == 'under' else actual >= target

        if mkt == 'HR':
            b = box.get('batters', {}).get(nm)
            if b is not None:
                win = b['hr'] >= 1; got = 'HR' if win else '0 HR'
            elif homers:
                win = nm in homers; got = 'HR' if win else '0 HR'
        elif mkt == 'TOTAL':
            tot = box.get('totals', {}).get(gamekey(p.get('game', '')))
            if tot is not None and p.get('ref_line'):
                line = float(p.get('ref_line')); lean = (p.get('lean') or 'OVER').upper()
                if tot == line:
                    got = f'{tot} runs · push'          # exact push -> void (stays ungraded)
                else:
                    over = tot > line
                    win = over if lean == 'OVER' else (not over)
                    got = f'{tot} runs'
        elif mkt == 'NRFI':
            fi = box.get('first_inning', {}).get(gamekey(p.get('game', '')))
            if fi is not None:
                scored = fi >= 1; lean = (p.get('lean') or 'NRFI').upper()
                win = (not scored) if lean == 'NRFI' else scored
                got = '0 in 1st' if not scored else f'{fi} in 1st'
        elif box:
            b = box['batters'].get(nm, {}); pi = box['pitchers'].get(nm, {})
            if mkt == 'HIT' and b: win = b['h'] >= 1; got = f'{b["h"]} H'
            elif mkt == 'HRR' and b:
                s = b['h']+b['r']+b['rbi']; win = s >= 1; got = f'{s} H+R+RBI'
                detail['actual'] = f'{b["h"]}H · {b["r"]}R · {b["rbi"]}RBI'
            elif mkt == 'SB' and b: win = b['sb'] >= 1; got = f'{b["sb"]} SB'
            elif mkt == '2B' and b: win = b['d'] >= 1; got = f'{b["d"]} 2B'
            elif mkt == 'TB' and b:
                tb = b['h'] + b['d'] + (2 * b.get('t', 0)) + (3 * b['hr'])
                win = tb >= p.get('win_at', 99); got = f'{tb} TB'
            elif mkt == 'K' and pi:
                win = pi['k'] >= p.get('win_at', 99); got = f'{pi["k"]} K'
                detail['actual'] = f'{pi["h"]}H · {pi["er"]}ER · {pi["outs"]} outs'
            elif mkt == 'OUTS' and pi:
                win = pi['outs'] >= p.get('win_at', 99); got = f'{pi["outs"]} outs'
                detail['actual'] = f'{pi["h"]}H · {pi["er"]}ER · {pi["outs"]} outs'
            elif mkt == 'H_ALLOWED' and pi:
                win = pi['h'] >= p.get('win_at', 99); got = f'{pi["h"]} H allowed'
                detail['actual'] = f'{pi["h"]}H · {pi["er"]}ER · {pi["outs"]} outs'
            elif mkt == 'OUTS_ALT' and pi:
                win = directional(pi['outs']); got = f'{pi["outs"]} outs'
                detail['actual'] = f'{pi["h"]}H · {pi["er"]}ER · {pi["outs"]} outs'
            elif mkt == 'H_ALLOWED_ALT' and pi:
                win = directional(pi['h']); got = f'{pi["h"]} H allowed'
                detail['actual'] = f'{pi["h"]}H · {pi["er"]}ER · {pi["outs"]} outs'
            elif mkt == 'ER_ALLOWED' and pi:
                win = pi['er'] >= p.get('win_at', 99); got = f'{pi["er"]} ER'
                detail['actual'] = f'{pi["h"]}H · {pi["er"]}ER · {pi["outs"]} outs'
        # K projected peripherals (the "added context") — shown graded or not
        ctx = p.get('context') or {}
        if mkt == 'K' and ctx:
            detail['proj'] = (f'proj {ctx.get("proj_hits_allowed","?")}H · '
                              f'{ctx.get("proj_runs_allowed","?")}R · {ctx.get("proj_era","?")} ERA')
        disp_name = p.get('name') or p.get('game') or p.get('pick', '')
        disp_line = p.get('line', '')
        if mkt == 'TOTAL':
            disp_line = f"{(p.get('lean') or '').title()} {p.get('ref_line', '')}".strip()
        elif mkt == 'NRFI':
            disp_line = (p.get('lean') or 'NRFI')
        graded.append({'market': mkt, 'name': disp_name, 'line': disp_line,
                       'consensus': p.get('consensus', 0), 'max': p.get('consensus_max', 6),
                       'win': win, 'got': got, 'pick': p.get('pick', ''), 'detail': detail,
                       'projection': p.get('projection'), 'main_line': p.get('main_line'),
                       'direction': p.get('direction'), 'alt_margin': p.get('alt_margin')})

    # ---- HR consensus buckets (real where graded) ----
    hr = [g for g in graded if g['market'] == 'HR' and g['win'] is not None]
    def band(lo, hi):
        sub = [g for g in hr if lo <= g['consensus'] <= hi]
        return sum(1 for g in sub if g['win']), sum(1 for g in sub if not g['win'])
    hr_buckets = []
    if hr:
        for label, lo, hi, cls in [('🔒 5–6 lenses', 5, 6, 'e'), ('4 lenses', 4, 4, 'e'),
                                   ('2–3 lenses', 2, 3, 's'), ('0–1 lenses', 0, 1, 'f')]:
            w, l = band(lo, hi)
            if w + l: hr_buckets.append({'name': label, 'cls': cls, 'w': w, 'l': l})
    hr_w = sum(1 for g in hr if g['win']); hr_l = sum(1 for g in hr if not g['win'])

    # ---- grades list (HR, by consensus desc) ----
    rows = sorted(hr, key=lambda g: -g['consensus'])[:14]
    grades = [{'wl': ('w' if g['win'] else 'l'), 'name': f'{g["name"]} — 1+ HR',
               'tag': f'{g["consensus"]}/{g["max"]} consensus · HR Board', 'got': g['got']}
              for g in rows]

    # ---- accumulate season history ----
    prev = load_json(RESULTS_FILE, {})
    hist = prev.get('history', [])
    hist = [h for h in hist if h.get('date') != date]  # idempotent per date
    hist.append({'date': date, 'hr_w': hr_w, 'hr_l': hr_l})
    season_w = sum(h['hr_w'] for h in hist); season_l = sum(h['hr_l'] for h in hist)
    last7 = hist[-7:]; l7w = sum(h['hr_w'] for h in last7); l7l = sum(h['hr_l'] for h in last7)
    trend, tlabels = [], []
    for h in hist[-14:]:
        t = h['hr_w'] + h['hr_l']
        trend.append(round(100*h['hr_w']/t) if t else 0); tlabels.append(h['date'])

    other_graded = any(g['win'] is not None for g in graded if g['market'] != 'HR')

    # ---- per-market summary (for the markets grid) ----
    META = [('HR','HR','🏆'),('K','K','⚡'),('OUTS','Outs','📏'),('HIT','Hits','🎯'),('HRR','HRR','💥'),
            ('TB','Total Bases','📏'),('H_ALLOWED','Hits Allowed','🚦'),('ER_ALLOWED','ER Allowed','📈'),
            ('H_ALLOWED_ALT','Hits Allowed Alt','🚦'),('OUTS_ALT','Outs Alt','📏'),
            ('TOTAL','Totals','📈'),('NRFI','NRFI','🥶'),('SB','SB','🏃'),('2B','2B','💎')]
    markets = []
    for key, label, icon in META:
        sub = [g for g in graded if g['market'] == key and g['win'] is not None]
        w = sum(1 for g in sub if g['win']); l = sum(1 for g in sub if not g['win'])
        markets.append({'key': key, 'label': label, 'icon': icon, 'w': w, 'l': l,
                        'graded': bool(sub),
                        'picks': sum(1 for g in graded if g['market'] == key)})

    # ---- per-market detail (rows + context) for the tab panels ----
    market_detail = {}
    for key, label, icon in META:
        sub = [g for g in graded if g['market'] == key]
        gd = [g for g in sub if g['win'] is not None]
        pend = [g for g in sub if g['win'] is None]
        gd_sorted = sorted(gd, key=lambda g: (-g['consensus'], 0 if g['win'] else 1))
        pend_sorted = sorted(pend, key=lambda g: -g['consensus'])
        rows = []
        for g in (gd_sorted + pend_sorted)[:18]:
            rows.append({
                'wl': ('w' if g['win'] else 'l') if g['win'] is not None else 'p',
                'name': g['name'], 'line': g['line'],
                'consensus': g['consensus'], 'max': g['max'],
                'got': g['got'], 'detail': g['detail'],
            })
        w = sum(1 for g in gd if g['win']); l = sum(1 for g in gd if not g['win'])
        market_detail[key] = {'label': label, 'icon': icon, 'w': w, 'l': l,
                              'graded': bool(gd), 'picks': len(sub), 'rows': rows}

    results = {
        'is_preview': not other_graded,
        'season_day': prev.get('season_day', 80),
        'updated': f'{date} graded',
        'season': {'w': season_w, 'l': season_l},
        'last7': {'w': l7w, 'l': l7l},
        'yesterday_date': date, 'yesterday': {'w': hr_w, 'l': hr_l},
        'markets': markets,
        'market_detail': market_detail,
        'k_buckets': [],  # filled when API grading runs
        'hr_buckets': hr_buckets,
        'hr_insight': ('not enough graded days yet to compare lens buckets; check back as '
                       'the sample builds.'),
        'trend': trend, 'trend_labels': tlabels,
        'grades': grades, 'history': hist,
    }
    with open(RESULTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=1)
    print(f'Graded {date}: HR board {hr_w}-{hr_l} '
          f'({"+API" if other_graded else "HR only; other markets pending"}). Wrote {RESULTS_FILE}')

if __name__ == '__main__':
    try:
        grade()
    except Exception as e:
        print('grade_results error (non-fatal):', e)
    sys.exit(0)
